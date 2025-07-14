#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import xml.etree.ElementTree as ET
import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import os
import re
from datetime import datetime
import traceback

class IBMXMLParser:
    def __init__(self):
        self.systems = []  # 시스템 단위로 저장
        
    def parse_cfblob(self, cfblob_text):
        """CFReportBLOB 텍스트 파싱"""
        lines = cfblob_text.strip().split('\n')
        current_system = None
        current_product = None
        
        for line in lines:
            if not line.strip():
                continue
                
            # 07로 시작: 시스템 헤더
            if line.startswith('07'):
                if current_system:
                    if current_product:
                        current_system['products'].append(current_product)
                    self.systems.append(current_system)
                
                # 시스템 이름 추출
                system_name = line[5:50].strip()
                current_system = {
                    'name': system_name,
                    'products': []
                }
                current_product = None
                
            # 08로 시작: 메인 제품 정보
            elif line.startswith('08'):
                if current_product and current_system:
                    current_system['products'].append(current_product)
                
                # 모델 번호 추출 (position 2-11)
                model_no = line[2:12].strip()
                # 하드웨어: 4657 924 -> 4657-924 형태로 변환
                if len(model_no) > 4 and model_no[4] == ' ':
                    formatted_model_no = model_no[:4] + '-' + model_no[5:]
                else:
                    formatted_model_no = model_no
                    
                current_product = {
                    'model_no': model_no,  # 원본 모델 번호 (매칭용)
                    'display_model_no': formatted_model_no,  # 표시용 모델 번호
                    'description': '',
                    'main_qty': 1,
                    'unit_price': 0,
                    'type': 'Hardware',  # 기본값
                    'subitems': []
                }
                
            # 95로 시작: 메인 제품 설명
            elif line.startswith('95') and current_product:
                # 설명 추출 (position 93부터)
                description = line[92:].strip()
                current_product['description'] = description
                
            # 96으로 시작: 서브 아이템
            elif line.startswith('96') and current_product:
                # 서브아이템 코드 (position 2-6)
                sub_code = line[2:6].strip()
                # 설명 (position 51부터)
                sub_desc = line[50:].strip()
                
                # N/C 체크
                is_nc = 'N' in line[25:30] if len(line) > 30 else False
                
                subitem = {
                    'code': sub_code,
                    'description': sub_desc,
                    'is_nc': is_nc,
                    'qty': 1,
                    'unit_price': 0,
                    'type': 'Hardware'  # 기본값
                }
                current_product['subitems'].append(subitem)
            
            # 47로 시작: 소프트웨어 제품 정보 (5692A6P, 5765G98 등)
            elif line.startswith('47'):
                if current_system:
                    # 4번째 문자부터 7글자가 모델 번호
                    sw_model_no = line[2:12].strip()
                    
                    # 소프트웨어 모델 번호의 하이픈 포맷팅
                    # 7자리인 경우: 5692A6P -> 5692-A6P, 5765G98 -> 5765-G98
                    # 8자리인 경우: 5773SM3 -> 5773-SM3, 5773RS3 -> 5773-RS3
                    if len(sw_model_no) >= 7:
                        # 첫 4자리 다음에 하이픈 추가
                        formatted_model_no = sw_model_no[:4] + '-' + sw_model_no[4:]
                    else:
                        formatted_model_no = sw_model_no
                    
                    sw_product = {
                        'model_no': sw_model_no,
                        'display_model_no': formatted_model_no,
                        'description': '',
                        'main_qty': 1,
                        'unit_price': 0,
                        'type': 'Software',
                        'subitems': []
                    }
                    current_product = sw_product
                
            # 95로 시작: 메인 제품 설명
            elif line.startswith('95') and current_product:
                # 설명 추출 (position 93부터)
                description = line[92:].strip()
                current_product['description'] = description
                
            # 96으로 시작: 서브 아이템
            elif line.startswith('96') and current_product:
                # 서브아이템 코드 (position 2-6)
                sub_code = line[2:6].strip()
                # 설명 (position 51부터)
                sub_desc = line[50:].strip()
                
                # N/C 체크
                is_nc = 'N' in line[25:30] if len(line) > 30 else False
                
                subitem = {
                    'code': sub_code,
                    'description': sub_desc,
                    'is_nc': is_nc,
                    'qty': 1,
                    'unit_price': 0
                }
                current_product['subitems'].append(subitem)
        
        # 마지막 시스템 추가
        if current_system:
            if current_product:
                current_system['products'].append(current_product)
            self.systems.append(current_system)
    
    def parse_xml_file(self, file_path):
        """XML 파일 파싱"""
        try:
            tree = ET.parse(file_path)
            root = tree.getroot()
            
            # CFReportBLOB 찾기
            cfblob = root.find('.//CFReportBLOB')
            if cfblob is not None and cfblob.text:
                self.parse_cfblob(cfblob.text)
            
            # XML에서 가격 정보 추출
            line_items = root.findall('.//ProductLineItem')
            
            for line_item in line_items:
                # 모델 번호 찾기
                prop_id = line_item.find('.//ProprietaryProductIdentifier')
                if prop_id is None:
                    continue
                    
                # 다양한 형태의 모델 번호 처리
                xml_model_no = prop_id.text
                normalized_model_no = xml_model_no.replace('-', '').replace(' ', '')
                
                # 제품 타입 찾기 (Hardware/Software)
                type_elem = line_item.find('.//ProductTypeCode')
                product_type = type_elem.text if type_elem is not None else 'Hardware'
                
                # 수량 찾기
                qty_elem = line_item.find('Quantity')
                qty = int(qty_elem.text) if qty_elem is not None else 1
                
                # 가격 찾기
                price_elem = line_item.find('.//MonetaryAmount')
                price = 0
                if price_elem is not None:
                    price_text = price_elem.text.replace(',', '').replace('N/C', '0')
                    try:
                        price = float(price_text)
                    except:
                        price = 0
                
                # 모든 시스템에서 매칭되는 제품 찾기
                for system in self.systems:
                    for product in system['products']:
                        # 제품 모델 번호도 정규화하여 비교
                        product_model_no = product['model_no'].replace(' ', '').replace('-', '')
                        if product_model_no == normalized_model_no:
                            product['main_qty'] = qty
                            product['unit_price'] = price
                            product['type'] = product_type  # 타입 업데이트
                            
                            # Software인 경우 display_model_no 재포맷팅
                            if product_type == 'Software' and len(product['model_no']) >= 7:
                                product['display_model_no'] = product['model_no'][:4] + '-' + product['model_no'][4:]
                            break
                        # 원본 모델 번호로도 비교
                        elif product['model_no'] == xml_model_no:
                            product['main_qty'] = qty
                            product['unit_price'] = price
                            product['type'] = product_type
                            
                            # Software인 경우 display_model_no 재포맷팅
                            if product_type == 'Software' and len(product['model_no']) >= 7:
                                product['display_model_no'] = product['model_no'][:4] + '-' + product['model_no'][4:]
                            break
            
            # 서브아이템 가격 정보 추출
            for line_item in line_items:
                subitems = line_item.findall('.//ProductSubLineItem')
                for subitem in subitems:
                    code_elem = subitem.find('.//ProprietaryProductIdentifier')
                    if code_elem is None:
                        continue
                    
                    code = code_elem.text
                    
                    # 제품 타입 찾기
                    type_elem = subitem.find('.//ProductTypeCode')
                    sub_type = type_elem.text if type_elem is not None else 'Hardware'
                    
                    # 수량
                    qty_elem = subitem.find('Quantity')
                    qty = int(qty_elem.text) if qty_elem is not None else 1
                    
                    # 가격
                    price_elem = subitem.find('.//MonetaryAmount')
                    price = 0
                    if price_elem is not None:
                        price_text = price_elem.text.replace(',', '').replace('N/C', '0')
                        try:
                            price = float(price_text)
                        except:
                            price = 0
                    
                    # 매칭되는 서브아이템 찾기
                    for system in self.systems:
                        for product in system['products']:
                            for sub in product['subitems']:
                                if sub['code'] == code:
                                    sub['qty'] = qty
                                    sub['unit_price'] = price
                                    sub['type'] = sub_type  # 타입 업데이트
                                    break
        
        except Exception as e:
            raise Exception(f"XML 파싱 오류: {str(e)}")

class XMLToExcelGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("IBM XML to Excel Converter")
        self.root.geometry("600x300")
        
        # 메인 프레임
        main_frame = ttk.Frame(root, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # 파일 선택 섹션
        file_frame = ttk.LabelFrame(main_frame, text="파일 선택", padding="10")
        file_frame.grid(row=0, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=5)
        
        self.file_path = tk.StringVar()
        ttk.Entry(file_frame, textvariable=self.file_path, width=60).grid(row=0, column=0, padx=5)
        ttk.Button(file_frame, text="Browse", command=self.browse_file).grid(row=0, column=1, padx=5)
        
        # 버튼 섹션
        button_frame = ttk.Frame(main_frame)
        button_frame.grid(row=1, column=0, columnspan=2, pady=20)
        
        ttk.Button(button_frame, text="변환", command=self.convert, width=20).grid(row=0, column=0, padx=5)
        
        # Status bar
        self.status = tk.StringVar()
        self.status.set("준비됨")
        ttk.Label(main_frame, textvariable=self.status).grid(row=2, column=0, columnspan=2, sticky=(tk.W, tk.E))
        
        # Configure grid weights
        root.columnconfigure(0, weight=1)
        root.rowconfigure(0, weight=1)
        main_frame.columnconfigure(0, weight=1)
    
    def browse_file(self):
        filename = filedialog.askopenfilename(
            title="XML 파일 선택",
            filetypes=[("XML files", "*.xml"), ("All files", "*.*")]
        )
        if filename:
            self.file_path.set(filename)
    
    def convert(self):
        if not self.file_path.get():
            messagebox.showwarning("경고", "파일을 선택해주세요.")
            return
        
        try:
            parser = IBMXMLParser()
            parser.parse_xml_file(self.file_path.get())
            
            # Get save location
            save_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                initialfile=f"IBM_Config_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            )
            
            if save_path:
                # Create Excel writer
                with pd.ExcelWriter(save_path, engine='openpyxl') as writer:
                    # Import styles
                    from openpyxl.styles import Alignment, Font, PatternFill
                    
                    # First, create TOTAL sheet
                    total_rows = []
                    total_excel_row = 2  # Start from row 2 (after header)
                    
                    # Add header for TOTAL sheet
                    total_rows.append({
                        'SYSTEM': 'SYSTEM',
                        'ProprietaryProductIdentifier': 'ProprietaryProductIdentifier',
                        'DESCRIPTION': 'DESCRIPTION',
                        "Q'TY": "Q'TY",
                        'UNIT PRICE': 'UNIT PRICE',
                        'EXTENDED PRICE': 'EXTENDED PRICE'
                    })
                    
                    # Collect all data across systems (only main products with 7-digit identifiers)
                    for system in parser.systems:
                        system_name = system['name']
                        
                        # Add only main products (not subitems)
                        for product in system['products']:
                            # Main product row
                            total_rows.append({
                                'SYSTEM': system_name,
                                'ProprietaryProductIdentifier': product.get('display_model_no', product['model_no']),
                                'DESCRIPTION': product['description'],
                                "Q'TY": product['main_qty'],
                                'UNIT PRICE': product['unit_price'],
                                'EXTENDED PRICE': f'=D{total_excel_row}*E{total_excel_row}'  # Excel formula
                            })
                            total_excel_row += 1
                    
                    # Create TOTAL DataFrame
                    total_df = pd.DataFrame(total_rows)
                    
                    # Write TOTAL sheet
                    total_df.to_excel(writer, index=False, sheet_name='TOTAL', header=False)
                    
                    # Format TOTAL sheet
                    total_worksheet = writer.sheets['TOTAL']
                    
                    # Format headers
                    for col in range(1, 7):  # Columns A-F
                        cell = total_worksheet.cell(row=1, column=col)
                        cell.font = Font(bold=True)
                        cell.alignment = Alignment(horizontal='center')
                        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
                    
                    # Format data rows
                    for row in range(2, len(total_df) + 1):
                        # System column - bold if not empty
                        system_cell = total_worksheet[f'A{row}']
                        if system_cell.value and system_cell.value not in ['', 'SYSTEM']:
                            system_cell.font = Font(bold=True)
                            system_cell.fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
                        
                        # Format price columns
                        for col in ['E', 'F']:  # UNIT PRICE, EXTENDED PRICE
                            cell = total_worksheet[f'{col}{row}']
                            if col == 'E' and isinstance(cell.value, (int, float)):
                                cell.number_format = '#,##0.00'
                            elif col == 'F':
                                cell.number_format = '#,##0.00'
                    
                    # Adjust column widths for TOTAL sheet
                    total_column_widths = {
                        'A': 30,  # SYSTEM
                        'B': 25,  # ProprietaryProductIdentifier
                        'C': 50,  # DESCRIPTION
                        'D': 8,   # Q'TY
                        'E': 15,  # UNIT PRICE
                        'F': 15   # EXTENDED PRICE
                    }
                    
                    for col, width in total_column_widths.items():
                        total_worksheet.column_dimensions[col].width = width
                    
                    # Then, create separate sheet for each system
                    for system in parser.systems:
                        # Extract system-specific data
                        system_name = system['name']
                        
                        # Create safe sheet name (Excel limits: 31 chars, no special chars)
                        safe_sheet_name = re.sub(r'[\\/*?:\[\]]', '_', system_name)[:31]
                        
                        # Build system-specific dataframe
                        rows = []
                        excel_row = 3  # Start from row 3 (after system name and header)
                        
                        # Add system name row (will be in row 1)
                        rows.append({
                            'TYPE': system_name,
                            'MODEL NO.': '',
                            'DESCRIPTION': '',
                            "Q'TY": '',
                            'UNIT PRICE': '',
                            'EXTENDED PRICE': ''
                        })
                        
                        # Add empty row (will be in row 2)
                        rows.append({
                            'TYPE': '',
                            'MODEL NO.': '',
                            'DESCRIPTION': '',
                            "Q'TY": '',
                            'UNIT PRICE': '',
                            'EXTENDED PRICE': ''
                        })
                        
                        # Add headers (will be in row 3)
                        rows.append({
                            'TYPE': 'TYPE',
                            'MODEL NO.': 'MODEL NO.',
                            'DESCRIPTION': 'DESCRIPTION',
                            "Q'TY": "Q'TY",
                            'UNIT PRICE': 'UNIT PRICE',
                            'EXTENDED PRICE': 'EXTENDED PRICE'
                        })
                        
                        excel_row = 4  # Data starts from row 4
                        
                        # Add products and subitems
                        for product in system['products']:
                            # Main product row
                            rows.append({
                                'TYPE': product.get('type', 'Hardware'),
                                'MODEL NO.': product.get('display_model_no', product['model_no']),
                                'DESCRIPTION': product['description'],
                                "Q'TY": product['main_qty'],
                                'UNIT PRICE': product['unit_price'],
                                'EXTENDED PRICE': f'=D{excel_row}*E{excel_row}'  # Excel formula
                            })
                            excel_row += 1
                            
                            # Subitem rows
                            for sub in product['subitems']:
                                rows.append({
                                    'TYPE': sub.get('type', 'Hardware'),
                                    'MODEL NO.': f"  {sub['code']}",  # Indent subitems
                                    'DESCRIPTION': f"  {sub['description']}",
                                    "Q'TY": sub['qty'],
                                    'UNIT PRICE': sub['unit_price'],
                                    'EXTENDED PRICE': f'=D{excel_row}*E{excel_row}'  # Excel formula
                                })
                                excel_row += 1
                        
                        # Create DataFrame
                        df = pd.DataFrame(rows)
                        
                        # Write to Excel
                        df.to_excel(writer, index=False, sheet_name=safe_sheet_name, header=False)
                        
                        # Format the sheet
                        worksheet = writer.sheets[safe_sheet_name]
                        
                        # Merge cells for system name
                        worksheet.merge_cells('A1:F1')
                        
                        # Format system name
                        system_cell = worksheet['A1']
                        system_cell.font = Font(bold=True, size=14)
                        system_cell.alignment = Alignment(horizontal='center', vertical='center')
                        system_cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
                        
                        # Format headers
                        for col in range(1, 7):  # Columns A-F
                            cell = worksheet.cell(row=3, column=col)
                            cell.font = Font(bold=True)
                            cell.alignment = Alignment(horizontal='center')
                            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
                        
                        # Format data rows
                        for row in range(4, len(df) + 1):
                            # Format price columns
                            for col in ['E', 'F']:  # UNIT PRICE, EXTENDED PRICE
                                cell = worksheet[f'{col}{row}']
                                if col == 'E' and isinstance(cell.value, (int, float)):
                                    cell.number_format = '#,##0.00'
                                elif col == 'F':
                                    cell.number_format = '#,##0.00'
                        
                        # Adjust column widths
                        column_widths = {
                            'A': 10,  # TYPE
                            'B': 15,  # MODEL NO.
                            'C': 50,  # DESCRIPTION  
                            'D': 8,   # Q'TY
                            'E': 15,  # UNIT PRICE
                            'F': 15   # EXTENDED PRICE
                        }
                        
                        for col, width in column_widths.items():
                            worksheet.column_dimensions[col].width = width
                
                messagebox.showinfo("완료", f"파일이 저장되었습니다:\n{save_path}")
                self.status.set("변환 완료")
            
        except Exception as e:
            error_msg = f"변환 중 오류 발생:\n{str(e)}\n\n{traceback.format_exc()}"
            messagebox.showerror("오류", error_msg)
            self.status.set("오류 발생")

def main():
    root = tk.Tk()
    app = XMLToExcelGUI(root)
    root.mainloop()

if __name__ == "__main__":
    main()